import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
for i in range(2, sheet.max_row + 1):
    cellobj = sheet.cell(row=i, column=2).value
    produce = sheet.cell(row=i, column=1).value
    if produce == 'Celery':
        cellobj = 1.19
    elif produce == 'Garlic':
        cellobj = 3.07
    elif produce == 'Lemon':
        cellobj = 1.27
    sheet.cell(row=i, column=2).value = cellobj
wb.save('UpdatedSheet.xlsx')