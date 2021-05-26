import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']

for i in range(1, 4):
    print('Column:', i, sheet.cell(row=1, column=i).value)