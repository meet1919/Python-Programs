import openpyxl, smtplib

wb = openpyxl.open('duesRecords.xlsx')
sheet = wb['Sheet1']

# Creating New Sheet for Pending
wb1 = openpyxl.Workbook()
sheet1 = wb1['Sheet']
sheet1['A1'] = 'Members'
sheet1['B1'] = 'Email Address'
member = []
email = []

for i in range(2, 8):
    if sheet.cell(row=i, column=8).value == None:
        member.append(sheet.cell(row=i, column=1).value)
        email.append(sheet.cell(row=i, column=2).value)
wb.close()

# Adding values to new sheet
for j in range(len(member)):
    sheet1.cell(row=j+2, column=1).value = member[j]
    sheet1.cell(row=j+2, column=2).value = email[j]
wb1.save('Pending-List.xlsx')

# Sending Mail
smtpobj = smtplib.SMTP('smtp.gmail.com', 587)
smtpobj.starttls()
smtpobj.login('meetgondaliya1999@gmail.com', 'Nigga@710')
for k in range(len(member)):
    msg = 'Subject:Reminder\nDear %s this is an reminder.\nKindly Pay your DUES.' %(member[k])
    print('Sending mail to ', email[k])
    smtpobj.sendmail('meetgondaliya1999@gmail.com', email[k], msg)
smtpobj.quit()